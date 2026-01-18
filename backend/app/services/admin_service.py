from app.services.sheets_service import sheets_service
from app.utils import validate_datetime, validate_question_type
from openpyxl import load_workbook
from io import BytesIO

class AdminService:
    @staticmethod
    def get_all_surveys():
        """Get all surveys"""
        return sheets_service.get_all_surveys()

    @staticmethod
    def create_survey(title, description, study_content_html, start_time, end_time,
                     duration_minutes, total_questions, pass_score, max_attempts=3):
        """Create a new survey"""
        if not title:
            raise ValueError('标题不能为空')
        if not validate_datetime(start_time) or not validate_datetime(end_time):
            raise ValueError('时间格式不正确')
        if start_time >= end_time:
            raise ValueError('开始时间不能晚于结束时间')
        return sheets_service.create_survey(
            title, description, study_content_html, start_time, end_time,
            duration_minutes, total_questions, pass_score, max_attempts
        )

    @staticmethod
    def update_survey(survey_id, title, description, study_content_html, start_time, end_time,
                     duration_minutes, total_questions, pass_score, max_attempts=3):
        """Update an existing survey"""
        if not title:
            raise ValueError('标题不能为空')
        if not validate_datetime(start_time) or not validate_datetime(end_time):
            raise ValueError('时间格式不正确')
        if start_time >= end_time:
            raise ValueError('开始时间不能晚于结束时间')
        return sheets_service.update_survey(
            survey_id, title, description, study_content_html, start_time, end_time,
            duration_minutes, total_questions, pass_score, max_attempts
        )

    @staticmethod
    def delete_survey(survey_id):
        """Delete a survey"""
        if not survey_id:
            raise ValueError('问卷ID不能为空')
        return sheets_service.delete_survey(survey_id)

    @staticmethod
    def add_questions_to_survey(survey_id, questions):
        """Add questions to a survey"""
        if not questions:
            raise ValueError('至少需要1道题目')
        for q in questions:
            if not q.get('question_text'):
                raise ValueError('题目内容不能为空')
            if not validate_question_type(q.get('question_type')):
                raise ValueError(f'无效题目类型: {q.get("question_type")}')
        return sheets_service.add_questions(survey_id, questions)

    @staticmethod
    def parse_excel_file(file):
        """Parse an Excel file and extract questions"""
        try:
            # Read the file into memory
            file_content = BytesIO(file.read())
            wb = load_workbook(file_content, read_only=True)
            ws = wb.active

            questions = []
            # Skip header row, start from row 2
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if not row or not row[0]:
                    continue

                # Column mapping:
                # A: question_type (single/multiple)
                # B: question_text
                # C-F: options (A-D)
                # G: correct_answer
                # H: score
                # I: explanation
                question_type = str(row[0]).strip().lower() if row[0] else 'single'
                question_text = str(row[1]).strip() if len(row) > 1 and row[1] else ''

                if not question_text:
                    continue

                # Parse options (columns C-F, indexes 2-5)
                options = []
                for i in range(2, 6):  # Columns C, D, E, F
                    if len(row) > i and row[i]:
                        options.append(str(row[i]).strip())

                if len(options) < 2:
                    raise ValueError(f'第{row_idx}行: 至少需要2个选项')

                # Parse correct answer
                correct_answer = str(row[6]).strip().upper() if len(row) > 6 and row[6] else 'A'

                # Validate correct answer
                if question_type == 'single':
                    if correct_answer not in ['A', 'B', 'C', 'D', 'E', 'F']:
                        raise ValueError(f'第{row_idx}行: 单选题答案必须是 A-F')
                else:
                    # Multiple choice: validate each answer
                    answers = [a.strip() for a in correct_answer.replace('，', ',').split(',')]
                    for a in answers:
                        if a not in ['A', 'B', 'C', 'D', 'E', 'F']:
                            raise ValueError(f'第{row_idx}行: 多选题答案必须是 A-F，用逗号分隔')
                    correct_answer = ','.join(answers)

                # Parse score
                score = 5
                if len(row) > 7 and row[7]:
                    try:
                        score = int(row[7])
                    except ValueError:
                        score = 5

                # Parse explanation
                explanation = str(row[8]).strip() if len(row) > 8 and row[8] else ''

                questions.append({
                    'question_type': question_type,
                    'question_text': question_text,
                    'options': options,
                    'correct_answer': correct_answer,
                    'score': score,
                    'explanation': explanation
                })

            if not questions:
                raise ValueError('Excel文件中没有找到有效题目')

            return questions
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f'解析Excel失败: {str(e)}')


admin_service = AdminService()
