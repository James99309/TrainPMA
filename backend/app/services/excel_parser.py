"""Excel 考卷解析服务"""
import json
from io import BytesIO
from openpyxl import load_workbook


class ExcelParser:
    """解析 Excel 考卷格式"""

    # 列定义（8列格式，与模板一致，已移除分值列）
    COLUMNS = {
        'type': 0,      # A: 题型 (single/multiple)
        'question': 1,  # B: 题目
        'optionA': 2,   # C: 选项A
        'optionB': 3,   # D: 选项B
        'optionC': 4,   # E: 选项C
        'optionD': 5,   # F: 选项D
        'answer': 6,    # G: 正确答案
        'explanation': 7  # H: 解析
    }

    # 题型映射
    TYPE_MAP = {
        'single': 'single_choice',
        'multiple': 'multiple_choice',
        '单选': 'single_choice',
        '多选': 'multiple_choice',
        '单选题': 'single_choice',
        '多选题': 'multiple_choice',
    }

    @staticmethod
    def parse(file_content: bytes) -> dict:
        """
        解析 Excel 文件

        Args:
            file_content: Excel 文件的二进制内容

        Returns:
            {
                'success': bool,
                'questions': list,
                'errors': list,
                'summary': dict
            }
        """
        errors = []
        questions = []

        try:
            workbook = load_workbook(filename=BytesIO(file_content), read_only=True)
            sheet = workbook.active

            # 跳过第一行标题
            rows = list(sheet.iter_rows(min_row=2, values_only=True))

            for row_idx, row in enumerate(rows, start=2):
                # 跳过空行
                if not row or not row[0]:
                    continue

                try:
                    question = ExcelParser._parse_row(row, row_idx)
                    if question:
                        questions.append(question)
                except ValueError as e:
                    errors.append(f"第{row_idx}行: {str(e)}")

            workbook.close()

            # 生成摘要（固定每题5分）
            summary = {
                'total': len(questions),
                'single_choice': len([q for q in questions if q['question_type'] == 'single_choice']),
                'multiple_choice': len([q for q in questions if q['question_type'] == 'multiple_choice']),
                'total_score': len(questions) * 5
            }

            return {
                'success': len(errors) == 0,
                'questions': questions,
                'errors': errors,
                'summary': summary
            }

        except Exception as e:
            return {
                'success': False,
                'questions': [],
                'errors': [f"解析文件失败: {str(e)}"],
                'summary': {}
            }

    @staticmethod
    def _parse_row(row: tuple, row_idx: int) -> dict:
        """解析单行数据（9列格式）"""
        cols = ExcelParser.COLUMNS
        option_letters = ['A', 'B', 'C', 'D']

        # 获取题型
        raw_type = str(row[cols['type']] or '').strip().lower()
        question_type = ExcelParser.TYPE_MAP.get(raw_type)
        if not question_type:
            raise ValueError(f"无效的题型: '{raw_type}'，支持: single, multiple, 单选, 多选")

        # 获取题目
        question_text = str(row[cols['question']] or '').strip()
        if not question_text:
            raise ValueError("题目内容不能为空")

        # 获取选项 (A-D)，同时检测实际有多少个选项
        options = []
        for i in range(4):  # 最多4个选项
            col_idx = cols['optionA'] + i
            if col_idx < len(row) and row[col_idx]:
                option_text = str(row[col_idx]).strip()
                if option_text:
                    options.append(option_text)

        if len(options) < 2:
            raise ValueError("至少需要2个选项")

        # 智能检测答案列位置
        # 方法：在选项之后查找第一个看起来像答案的列（A-D 或 A,B,C 格式）
        answer_col_idx = None
        raw_answer = None

        # 从选项结束后开始查找
        for check_idx in range(cols['optionA'] + len(options), min(len(row), cols['optionA'] + 8)):
            if row[check_idx]:
                cell_value = str(row[check_idx]).strip().upper()
                # 检查是否是有效的答案格式（单个字母或逗号分隔的字母）
                cleaned = cell_value.replace(' ', '').replace(',', '').replace('，', '')
                if cleaned and all(c in option_letters[:len(options)] for c in cleaned):
                    answer_col_idx = check_idx
                    raw_answer = cell_value
                    break

        # 如果智能检测失败，回退到固定列位置
        if answer_col_idx is None:
            # 尝试9列格式（答案在G列，索引6）
            if len(row) > 6 and row[6]:
                cell_value = str(row[6]).strip().upper()
                cleaned = cell_value.replace(' ', '').replace(',', '').replace('，', '')
                if cleaned and all(c in option_letters[:len(options)] for c in cleaned):
                    answer_col_idx = 6
                    raw_answer = cell_value

        # 如果还是找不到，使用默认列位置（答案在G列，索引6）
        if answer_col_idx is None:
            answer_col_idx = cols['answer']
            raw_answer = str(row[answer_col_idx] if answer_col_idx < len(row) else '').strip().upper()

        if not raw_answer:
            raise ValueError("正确答案不能为空")

        # 解析答案
        if question_type == 'multiple_choice':
            # 多选答案：A,B,D 或 ABD
            answer_parts = raw_answer.replace(' ', '').replace(',', '').replace('，', '')
            correct_answer = list(answer_parts)
            # 验证答案是否在选项范围内
            for ans in correct_answer:
                if ans not in option_letters[:len(options)]:
                    raise ValueError(f"答案 '{ans}' 不在选项范围内")
        else:
            # 单选答案
            correct_answer = raw_answer[0] if raw_answer else ''
            if correct_answer not in option_letters[:len(options)]:
                raise ValueError(f"答案 '{correct_answer}' 不在选项范围内")

        # 获取解析 (可选) - 答案列之后
        explanation_col_idx = answer_col_idx + 1
        explanation = str(row[explanation_col_idx] or '').strip() if explanation_col_idx < len(row) else ''

        return {
            'question_type': question_type,
            'question_text': question_text,
            'options': options,
            'correct_answer': correct_answer,
            'explanation': explanation
        }

    @staticmethod
    def generate_template() -> bytes:
        """生成 Excel 模板"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "考卷模板"

        # 设置标题行（已移除分值列，固定每题5分）
        headers = ['题型', '题目', '选项A', '选项B', '选项C', '选项D', '正确答案', '解析']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # 添加示例数据
        examples = [
            ['single', '以下哪个是 Python 的关键字？', 'class', 'Class', 'CLASS', 'CLAS', 'A', 'class 是 Python 的保留关键字'],
            ['multiple', '以下哪些是 Python 的数据类型？', 'int', 'str', 'func', 'bool', 'A,B,D', 'int、str、bool 都是 Python 的内置数据类型'],
            ['single', '1+1=?', '1', '2', '3', '4', 'B', '基础数学'],
        ]

        for row_idx, example in enumerate(examples, 2):
            for col_idx, value in enumerate(example, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 调整列宽
        column_widths = [10, 40, 20, 20, 20, 20, 12, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # 添加说明行
        ws.cell(row=6, column=1, value="说明：")
        ws.cell(row=7, column=1, value="1. 题型：single=单选，multiple=多选")
        ws.cell(row=8, column=1, value="2. 多选答案用逗号分隔：A,B,D")
        ws.cell(row=9, column=1, value="3. 每题固定5分")
        ws.cell(row=10, column=1, value="4. 解析为可选项")

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


excel_parser = ExcelParser()
